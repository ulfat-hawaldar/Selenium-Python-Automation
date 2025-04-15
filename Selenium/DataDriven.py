import time

from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By

from Selenium.Login import password

# Load the Excel Sheet

workbook = load_workbook('Testdata.xlsx')

# Selecting Active Sheet
sheet = workbook.active

driver = webdriver.Firefox()
driver.maximize_window()

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    Username = row[0]
    Password = row[1]

    driver.get("https://www.saucedemo.com/")
    time.sleep(5)
    driver.find_element(By.ID,"user-name").send_keys(Username)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.ID,"login-button").click()
    time.sleep(5)

    driver.find_element(By.ID,"react-burger-menu-btn").click()
    time.sleep(5)
    driver.find_element(By.XPATH,"//a[@id='logout_sidebar_link']").click()
    time.sleep(5)
driver.quit()
